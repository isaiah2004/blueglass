"""Read Murai's four workbooks into pericopes and literary structures.

Purpose
    Turn 1.6 MB of spreadsheet into the two row shapes ``Q-009`` and ``Q-015``
    need, dropping the third-party verse quotations on the way out.

Key responsibilities
    - Walk each sheet's rows as a small state machine: unit header, node rows,
      legend row, blank separator.
    - Resolve every span through the sheet's book table.
    - Count what was skipped, repaired and dropped, so the loader can assert on
      measurements rather than on hope.

Traps this file exists to absorb
    - 228 node rows sit under a header that is a bare span rather than ``[n]``.
      They are sub-analyses of part of a pericope, they have no pericope to
      attach to, and they are skipped and counted.
    - Two node labels are missing their opening bracket (``A519:23)``). They are
      repaired, and the repair is counted.
    - The English column is a licence hazard; see ``murai_copyright``.
    - The Japanese column is read NOWHERE in this file, on purpose.

Dependencies
    openpyxl, plus this package's span, book, pattern and copyright modules.

Usage
    pericopes = read_pericopes()
    units, tally = read_structures()
"""

from __future__ import annotations

import re

from openpyxl import load_workbook

from scripts.murai_books import MuraiSheet, require_sheet
from scripts.murai_copyright import safe_gloss
from scripts.murai_patterns import Shape, classify, is_primed, pair_label
from scripts.murai_records import (
    PERICOPE_LIST_FILES,
    STRUCTURE_FILES,
    ParseTally,
    Pericope,
    StructureNode,
    StructureUnit,
    passage_id_for,
)
from scripts.murai_spans import Span, SpanError, parse_span_with
from scripts.raw_datasets import MURAI_STRUCTURE, verify_digest

#: A unit header: the pericope's ordinal in square brackets.
_UNIT_HEADER = re.compile(r"^\[(\d+)\]$")

#: A node cell: a label followed by its span in brackets.
_NODE = re.compile(r"^(?P<label>[^()]{1,12}?)\((?P<span>[^)]*)\)$")

#: The same cell with its opening bracket missing -- two rows in the corpus.
#: The prime mark is written as an escape rather than as itself: the workbooks
#: use both the ASCII and the typographic apostrophe, and only one of them is
#: legible in a diff.
_NODE_REPAIR = re.compile(
    "^(?P<label>[A-Z][0-9]?['\\u2019]?)"
    r"(?P<span>\d{1,3}:\d{1,3}[a-d]?(?:-(?:\d{1,3}:)?\d{1,3}[a-d]?)?)\)$"
)

_CHAPTER_MODULUS = 1_000


def _cell(row: tuple[object, ...], index: int) -> object:
    """One cell, tolerating the short tuples openpyxl yields for ragged rows."""
    return row[index] if len(row) > index else None


def _text(row: tuple[object, ...], index: int) -> str:
    """One cell as stripped text; the empty string when it is blank."""
    value = _cell(row, index)
    return "" if value is None else str(value).strip()


def _is_blank(row: tuple[object, ...]) -> bool:
    """True for the separator rows between units."""
    return not any(str(value).strip() for value in row if value is not None)


def _pericope_from(row: tuple[object, ...], sheet: MuraiSheet) -> Pericope | None:
    """One list row into one Pericope, or None for a row that is not one."""
    ordinal = _cell(row, 0)
    span_text = _text(row, 1)
    if not isinstance(ordinal, int) or not span_text:
        return None
    span = parse_span_with(span_text, sheet.resolve)
    return Pericope(
        passage_id=passage_id_for(span.start_key, span.end_key),
        book_number=span.book_number,
        chapter=(span.start_key // _CHAPTER_MODULUS) % _CHAPTER_MODULUS,
        start_key=span.start_key,
        end_key=span.end_key,
        title=_text(row, 2),
        ordinal=ordinal,
    )


def read_pericopes() -> list[Pericope]:
    """Every pericope boundary in the canon, from the two list workbooks."""
    pericopes: list[Pericope] = []
    for name in PERICOPE_LIST_FILES:
        path = verify_digest(MURAI_STRUCTURE, name)
        workbook = load_workbook(path, read_only=True, data_only=True)
        try:
            for sheet_name in workbook.sheetnames:
                sheet = require_sheet(sheet_name)
                for row in workbook[sheet_name].iter_rows(values_only=True):
                    pericope = _pericope_from(row, sheet)
                    if pericope is not None:
                        pericopes.append(pericope)
        finally:
            workbook.close()
    return pericopes


class _UnitBuilder:
    """Accumulates one unit's rows until the blank line that closes it."""

    def __init__(self, span: Span) -> None:
        self.span = span
        self.labels: list[str] = []
        self.nodes: list[StructureNode] = []
        self.legend: str | None = None

    def add_node(
        self, label: str, start_key: int, end_key: int, row: tuple[object, ...]
    ) -> None:
        """Append one labelled limb, keeping only the safe English gloss."""
        self.labels.append(label)
        self.nodes.append(
            StructureNode(
                node_index=len(self.nodes),
                label=label,
                pair_label=pair_label(label),
                is_centre=False,
                start_key=start_key,
                end_key=end_key,
                summary=safe_gloss(_cell(row, 2)),
                catchword=(_text(row, 3) or None),
            )
        )

    def build(self, book_number: int) -> StructureUnit:
        """Freeze the unit, marking the pivot the labels imply."""
        shape: Shape = classify(self.labels)
        nodes = tuple(
            StructureNode(
                node_index=node.node_index,
                label=node.label,
                pair_label=node.pair_label,
                is_centre=(
                    shape.centre is not None
                    and node.pair_label == shape.centre
                    and not is_primed(node.label)
                ),
                start_key=node.start_key,
                end_key=node.end_key,
                summary=node.summary,
                catchword=node.catchword,
            )
            for node in self.nodes
        )
        return StructureUnit(
            passage_id=passage_id_for(self.span.start_key, self.span.end_key),
            book_number=book_number,
            span=self.span,
            pattern=shape.pattern,
            centre_label=shape.centre,
            legend=self.legend,
            nodes=nodes,
        )


def _match_node(cell_text: str, tally: ParseTally) -> re.Match[str] | None:
    """Parse a node cell, repairing the two known missing-bracket rows."""
    match = _NODE.fullmatch(cell_text)
    if match is not None:
        return match
    repaired = _NODE_REPAIR.fullmatch(cell_text)
    if repaired is not None:
        tally.repaired_labels += 1
    return repaired


def _absorb_node(
    match: re.Match[str],
    builder: _UnitBuilder | None,
    book_number: int,
    row: tuple[object, ...],
    tally: ParseTally,
) -> None:
    """Attach one node to the open unit, or count it as an orphan."""
    if builder is None:
        tally.orphan_nodes += 1
        return
    try:
        span = parse_span_with(match.group("span"), lambda _token: book_number)
    except (SpanError, ValueError):
        tally.skipped_rows += 1
        return
    builder.add_node(match.group("label").strip(), span.start_key, span.end_key, row)
    tally.nodes += 1
    if _text(row, 2):
        if builder.nodes[-1].summary is None:
            tally.glosses_dropped += 1
        else:
            tally.glosses_kept += 1


def _read_sheet(
    rows: list[tuple[object, ...]], sheet: MuraiSheet, tally: ParseTally
) -> list[StructureUnit]:
    """Walk one worksheet's rows, emitting a unit per ``[n]`` block."""
    units: list[StructureUnit] = []
    builder: _UnitBuilder | None = None
    book_number = sheet.default_book

    def close() -> None:
        nonlocal builder
        if builder is None:
            return
        if builder.nodes:
            units.append(builder.build(book_number))
        else:
            # Murai marks some pericopes as having no chiastic structure at all
            # -- a single description line and nothing else. Counting them is
            # the difference between "no structure here" and "the parser broke".
            tally.unstructured_units += 1
        builder = None

    for row in rows:
        first = _text(row, 0)
        if _is_blank(row):
            close()
            continue
        if _UNIT_HEADER.fullmatch(first):
            close()
            span = parse_span_with(_text(row, 1), sheet.resolve)
            book_number = span.book_number
            builder = _UnitBuilder(span)
            tally.units += 1
            continue
        match = _match_node(first, tally)
        if match is not None:
            _absorb_node(match, builder, book_number, row, tally)
            continue
        if not first and builder is not None and builder.nodes:
            builder.legend = safe_gloss(_cell(row, 2)) or builder.legend
            continue
        tally.skipped_rows += 1
    close()
    return units


def read_structures() -> tuple[list[StructureUnit], ParseTally]:
    """Every literary structure in the canon, plus what the parse cost."""
    units: list[StructureUnit] = []
    tally = ParseTally()
    for name in STRUCTURE_FILES:
        path = verify_digest(MURAI_STRUCTURE, name)
        workbook = load_workbook(path, read_only=True, data_only=True)
        try:
            for sheet_name in workbook.sheetnames:
                sheet = require_sheet(sheet_name)
                rows = list(workbook[sheet_name].iter_rows(values_only=True))
                units.extend(_read_sheet(rows, sheet, tally))
        finally:
            workbook.close()
    return units, tally
